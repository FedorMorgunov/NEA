from __future__ import division
import string
from random import randint
import random
from tkMessageBox import showinfo
from openpyxl import load_workbook
import tkinter as tk
from openpyxl.styles import Font
from tkinter import ttk
from PIL import Image, ImageTk
from tkcalendar import DateEntry
from matplotlib import pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from pandas import DataFrame
from datetime import date
import sys
import os
# importing all essential libraries

global current_user
global set_of_questions
global used_questions
global current_game_results
# all global variables in the program
# could be replaced with accessor methods


def unicode_to_ascii(text):
    text = (text.
            replace('\xe2\x80\x99', "'").
            replace('\xc3\xa9', 'e').
            replace('\xe2\x80\x90', '-').
            replace('\xe2\x80\x91', '-').
            replace('\xe2\x80\x92', '-').
            replace('\xe2\x80\x93', '-').
            replace('\xe2\x80\x94', '-').
            replace('\xe2\x80\x94', '-').
            replace('\xe2\x80\x98', "'").
            replace('\xe2\x80\x9b', "'").
            replace('\xe2\x80\x9c', '"').
            replace('\xe2\x80\x9c', '"').
            replace('\xe2\x80\x9d', '"').
            replace('\xe2\x80\x9e', '"').
            replace('\xe2\x80\x9f', '"').
            replace('\xe2\x80\xa6', '...').
            replace('\xe2\x80\xb2', "'").
            replace('\xe2\x80\xb3', "'").
            replace('\xe2\x80\xb4', "'").
            replace('\xe2\x80\xb5', "'").
            replace('\xe2\x80\xb6', "'").
            replace('\xe2\x80\xb7', "'").
            replace('\xe2\x81\xba', "+").
            replace('\xe2\x81\xbb', "-").
            replace('\xe2\x81\xbc', "=").
            replace('\xe2\x81\xbd', "(").
            replace('\xe2\x81\xbe', ")")
                 )
    return text
# Source: https://gist.github.com/tushortz/9fbde5d023c0a0204333267840b592f9


class User:

    def __init__(self, name):
        self.name = name
        self.frame = tk.Frame(root)

    def back_to_menu_btn(self):
        back_btn = tk.Button(self.frame, text="Back", command=lambda: [self.clear_frame(), self.set_up_menu_frame()])
        back_btn.pack(pady=15)

    def set_up_menu_frame(self):
        invite_code = tk.StringVar()

        workbook = load_workbook(filename="users.xlsx")

        def submit():
            teachers = []
            students = []
            usernames = []
            types = []

            sheet = workbook.active
            for row1 in sheet["A"]:
                usernames.append(row1.value)
            usernames.pop(0)
            for row2 in sheet["C"]:
                types.append(row2.value)
            types.pop(0)

            for i in range(len(types)-1):
                if types[i] == 'Teacher':
                    teachers.append(usernames[i])

            found = False

            while (found is not True) and (len(teachers) > 0):
                for i in teachers:
                    temp_worksheet = workbook[i + ".xlsx"]
                    if temp_worksheet['A2'].value == invite_code.get():
                        for row in temp_worksheet['B']:
                            if row.value is not None:
                                students.append(row.value)
                        temp_worksheet['B' + str(len(students)+1)].value = self.name
                        showinfo(
                            title='Information',
                            message="Congratulations! You've been added to the tutor group. Please log out to update "
                                    "your account status.")

                        i = 1
                        for row in sheet['A']:
                            if row.value != self.name:
                                i += 1
                            else:
                                sheet['C' + str(i)].value = 'Student'

                        sub_btn.pack_forget()
                        invite_code_entry.pack_forget()

                        workbook.save(filename="users.xlsx")
                        found = True
                    else:
                        teachers.remove(i)

        invite_code_entry = tk.Entry(self.frame, textvariable=invite_code, font=('calibre', 10, 'normal'))
        invite_code_entry.pack(side=tk.TOP, anchor=tk.NW)

        sub_btn = tk.Button(self.frame, text='Submit', command=submit)
        sub_btn.pack(side=tk.TOP, anchor=tk.NW)

        self.show_menu_template()

        self.frame.pack(fill='both', expand=1)

    def show_menu_template(self):
        self.show_stats()

        load = Image.open("/Users/fedor/Downloads/menu.png")
        w, h = load.size
        load = load.resize((w // 5, h // 5))
        render = ImageTk.PhotoImage(load)
        img = tk.Label(self.frame, image=render)
        img.image = render
        img.pack(side=tk.TOP)

        b1 = tk.Button(self.frame, cursor='heart', bd=0, bg='white', text='questions by topic',
                       command=self.set_up_new_game)
        b1.pack(side=tk.TOP)

    def hide_menu_frame(self):
        self.frame.forget()

    # clears the frame, so there is  no need to make a new one

    def clear_frame(self):
        for widget in self.frame.winfo_children():
            widget.destroy()

    def choose_options(self):
        options = [
            1,
            2,
            3,
            4,
            5,
            6,
            7,
            8,
            9,
            10,
            11,
        ]

        # datatype of menu text
        clicked = tk.IntVar()

        # initial menu text
        clicked.set(11)

        # Create Dropdown menu
        drop = tk.OptionMenu(self.frame, clicked, *options)
        drop.pack()

        options1 = [
            "Traffic signs",
            "Rules of the road"
        ]

        # datatype of menu text
        clicked1 = tk.StringVar()

        # initial menu text
        clicked1.set("Traffic signs")

        # Create Dropdown menu
        drop1 = tk.OptionMenu(self.frame, clicked1, *options1)
        drop1.pack()

        # Create button
        button = tk.Button(self.frame, text="Continue", command=lambda: self.start(clicked.get(), clicked1.get()))
        button.pack(pady=15)

        self.back_to_menu_btn()

    def start(self, clicked, clicked1):
        self.clear_frame()
        self.set_up_game(clicked, '_'.join(clicked1.lower().split(' ')))

    def set_up_new_game(self):
        self.clear_frame()
        self.choose_options()

    def set_up_game(self, num_of_questions, topic):

        global set_of_questions
        global used_questions
        global current_game_results

        set_of_questions = []
        current_game_results = []
        used_questions = []

        if topic == "traffic_signs":
            results_cell = 'A1'
        else:
            results_cell = 'A2'

        workbook = load_workbook(filename="users.xlsx")
        worksheet = workbook[self.name + ".xlsx"]

        if worksheet[results_cell].value is None:
            set_of_questions = create_set_of_questions(topic, num_of_questions, self.name)
            self.hide_menu_frame()
            set_of_questions[0].display_question_frame()

        else:
            a = worksheet[results_cell].value.split(',')
            for i in a:
                if i[0] == 'W':
                    set_of_questions.append(get_question_at(int(i[1:]), topic, self.name))

            if len(set_of_questions) > num_of_questions:
                while len(set_of_questions) != num_of_questions:
                    random.shuffle(set_of_questions)
                    del set_of_questions[0]

            elif len(set_of_questions) < num_of_questions:
                while len(set_of_questions) != num_of_questions:
                    temp_question = Question(topic, self.name)
                    if temp_question not in set_of_questions:
                        set_of_questions.append(temp_question)

            random.shuffle(set_of_questions)
            self.hide_menu_frame()
            set_of_questions[0].display_question_frame()

    def show_stats(self):
        workbook = load_workbook(filename="users.xlsx")
        worksheet = workbook[self.name + ".xlsx"]

        try:
            results = map(int, worksheet['A3'].value.split(', '))
            game_nums = list(range(1, len(results) + 1))

            data1 = {}
            data1.update({'GAME': game_nums})
            data1.update({'PERCENTAGE': results})

            df1 = DataFrame(data1, columns=['GAME', 'PERCENTAGE'])

            figure1 = plt.Figure(figsize=(6, 5), dpi=60)
            ax1 = figure1.add_subplot(111)
            bar1 = FigureCanvasTkAgg(figure1, self.frame)
            bar1.get_tk_widget().pack(side=tk.BOTTOM, anchor=tk.SE)
            df1 = df1[['GAME', 'PERCENTAGE']].groupby('GAME').sum()
            df1.plot(kind='bar', legend=True, ax=ax1)
            ax1.set_title('PERFORMANCE PROGRESS')

            # Code adapted from this source: https://datatofish.com/matplotlib-charts-tkinter-gui/

        except:
            pass


def diff_dates(date1, date2):
    return abs(date2 - date1).days

# function that returns difference between two dates in days


class Student(User):

    def set_up_menu_frame(self):
        self.show_menu_template()

        def show_tasks():

            self.clear_frame()
            step = 0
            workbook = load_workbook(filename="users.xlsx")
            worksheet = workbook[self.name + ".xlsx"]

            for cell in worksheet['B']:

                if cell.value is not None:
                    step += 1

                    if worksheet['C'+str(step)].value == 'Not complete':

                        components = cell.value.split('/')

                        num_of_questions = int(components[0])
                        topic = components[1]
                        deadline = map(int, components[2].split('.'))
                        # changing data type of each component in the list to int

                        d2 = date(deadline[2], deadline[1], deadline[0])
                        # formatting to date

                        today = date.today()
                        # current date

                        result = str(diff_dates(d2, today)) + " days"

                        if diff_dates(d2, today) == 1:
                            result = "tomorrow"

                        tk.Label(self.frame,
                                 text=str(num_of_questions) + " questions on " + topic + " is due in " + result).pack()

                        complete = tk.Button(self.frame, text='complete', command=lambda: [self.start(num_of_questions, topic), self.update_status(step)])
                        complete.pack()

                        self.back_to_menu_btn()
                        break

                    elif (worksheet['C'+str(step)].value == 'Completed') and (step == len(worksheet['C'])):
                        tk.Label(self.frame, text="No homework due in soon!").pack()
                        self.back_to_menu_btn()

                else:
                    tk.Label(self.frame, text="Your teacher hasn't set any work yet.").pack()
                    self.back_to_menu_btn()
                    break

            # display most recent task

        btn = tk.Button(self.frame, text='Tasks', command=show_tasks)
        btn.pack()

        self.frame.pack(fill='both', expand=1)

    def update_status(self, step):
        workbook = load_workbook(filename="users.xlsx")
        worksheet = workbook[self.name + ".xlsx"]
        worksheet['C' + str(step)] = "Completed"
        workbook.save(filename="users.xlsx")


class Teacher(User):

    def get_students(self):
        students = []
        workbook = load_workbook(filename="users.xlsx")
        worksheet = workbook[self.name + ".xlsx"]
        for row in worksheet['B']:
            if row.value is not None:
                students.append(row.value)
        students.pop(0)
        return students

    def get_invite_code(self):
        workbook = load_workbook(filename="users.xlsx")
        worksheet = workbook[self.name + ".xlsx"]

        if worksheet['A2'].value is None:
            invite_code = ''.join(random.SystemRandom().choice(string.ascii_uppercase + string.digits) for _ in range(5)
                                  )
            a1 = worksheet['A1']
            a1.value = "Invite code"
            a1.font = Font(bold=True)
            b1 = worksheet['B1']
            b1.value = "Students"
            b1.font = Font(bold=True)
            worksheet['A2'] = invite_code
            workbook.save(filename="users.xlsx")

        else:
            invite_code = worksheet['A2'].value

        return invite_code

    def set_up_menu_frame(self):
        tk.Label(self.frame, text="Your unique tutor group invite code is: " + self.get_invite_code())\
            .pack(side=tk.TOP, anchor=tk.NW)

        def set_assignment():
            self.clear_frame()
            tk.Label(self.frame, text="Choose the topic and number of questions:").pack(pady=15)
            self.choose_options()

        set_work = tk.Button(self.frame, text="set task", command=set_assignment)
        set_work.pack()

        self.frame.pack(fill='both', expand=1)

    def start(self, clicked, clicked1):

        def set_task(date):
            wb = load_workbook(filename="users.xlsx")

            for student in self.get_students():
                i = 1
                temp_ws = wb[student + ".xlsx"]

                for cell in temp_ws['B']:
                    if cell.value is not None:
                        i += 1

                temp_ws['B' + str(i)].value = str(clicked) + "/" + clicked1 + "/" + date
                temp_ws['C' + str(i)].value = "Not complete"
                wb.save("users.xlsx")

            self.clear_frame()
            self.set_up_menu_frame()
            showinfo(title='Information', message='Task set!')
            # return to menu

        def choose_deadline():
            tk.Label(self.frame, text="Choose the deadline:").pack(pady=15)

            sel = tk.StringVar()
            today = date.today()

            cal = DateEntry(self.frame, selectmode='day', mindate=today, textvariable=sel)
            cal.pack()

            # mindate validates the date so homework can only be set for the future dates

            btn = tk.Button(self.frame, text='Set', command=lambda: set_task(sel.get()))
            btn.pack(pady=15)

            back_btn = tk.Button(self.frame, text="Back",
                                 command=lambda: [self.clear_frame(), self.choose_options()])
            back_btn.pack(pady=15)

        self.clear_frame()
        choose_deadline()


class Authorisation:
    def __init__(self, type):
        self.frame = ttk.Frame(tc)
        self.type = type

    def login_clicked(self, password, login, ut):

        usernames = []
        passwords = []
        types = []

        workbook = load_workbook(filename="users.xlsx")
        sheet = workbook.active
        for row1 in sheet["A"]:
            usernames.append(row1.value)
        for row2 in sheet["B"]:
            passwords.append(row2.value)
        for row3 in sheet["C"]:
            types.append(row3.value)

        # login
        if self.type == "Login":
            checked = False
            for i in range(1, len(usernames)):
                if usernames[i] == login and passwords[i] == password:
                    checked = True
                    type = types[i]

            if checked:
                print("Logged in successfully!")
                tc.pack_forget()

                global current_user

                if type == 'Teacher':
                    current_user = Teacher(login)

                elif type == 'Student':
                    current_user = Student(login)

                else:
                    current_user = User(login)

                current_user.set_up_menu_frame()

            else:
                showinfo(
                    title='Information',
                    message='Incorrect details. Try again')

        # register
        elif self.type == "Register":

            if entrycheck(login) and entrycheck(password):

                if login not in usernames:
                    sheet["A" + str(len(usernames) + 1)] = login
                    sheet["B" + str(len(usernames) + 1)] = password
                    if ut == 1:
                        sheet["C" + str(len(usernames) + 1)] = 'Teacher'
                    else:
                        sheet["C" + str(len(usernames) + 1)] = 'Not specified'
                    workbook.create_sheet(login + ".xlsx")
                    workbook.save(filename="users.xlsx")

                    showinfo(
                        title='Information',
                        message='Account created successfully. You can now log in')

                else:
                    showinfo(
                        title='Information',
                        message='Username taken. Try another one or log into an existing account')

            else:
                showinfo(
                    title='Information',
                    message='Invalid input. Try again')

    def display_sign_in(self):
        i = tk.IntVar()
        username = tk.StringVar()
        password = tk.StringVar()

        username_label = ttk.Label(self.frame, text="Username:")
        username_label.pack(fill='x', expand=True)

        username_label = ttk.Entry(self.frame, textvariable=username)
        username_label.pack(fill='x', expand=True)
        username_label.focus()

        password_label = ttk.Label(self.frame, text="Password:")
        password_label.pack(fill='x', expand=True)

        password_entry = ttk.Entry(self.frame, textvariable=password, show="*")
        password_entry.pack(fill='x', expand=True)

        login_button = ttk.Button(self.frame, text=self.type, command=lambda: self.login_clicked(password_entry.get(),
                                                                                                 username_label.get(),
                                                                                                 i.get()))
        login_button.pack(fill='x', expand=True, pady=10)

        if self.type == "Register":
            c = tk.Checkbutton(self.frame, text="Teacher", variable=i)
            c.pack()

        self.frame.pack(padx=10, pady=10, fill='x', expand=True)


def save_game_result(current_game_results, type, name):

    w = 0
    r = 0

    for i in current_game_results:
        if i[0] == 'W':
            w += 1
        else:
            r += 1

    percentage = r/(r+w)*100

    workbook = load_workbook(filename="users.xlsx")
    worksheet = workbook[name + ".xlsx"]

    if type == "traffic_signs":
        worksheet['A1'] = ','.join(current_game_results)
    else:
        worksheet['A2'] = ','.join(current_game_results)

    if worksheet['A3'].value is not None:
        worksheet['A3'] = worksheet['A3'].value + ', ' + str(int(percentage))
    else:
        worksheet['A3'] = str(int(percentage))

    workbook.save(filename="users.xlsx")


def get_question_at(pos, type, name):

    temp = Question(type, name)
    while temp.id != pos:
        temp = Question(type, name)
    return temp
# function that returns a question at a specified question (topic  and number of question are sent as parameters)
# used when mapping out questions answered incorrectly and adding them to the next game


class Question:
    def __init__(self, type, name):
        self.id = randint(1,11)
        self.type = type
        self.question_data = self.load_question()
        self.image = self.load_image()
        self.frame = tk.Frame(root)
        self.name = name

    def load_image(self):
        try:
            return Image.open("/Users/fedor/Desktop/traffic_signs/" + str(self.id) + ".jpg")
        # some images are saved as '.png' file
        except IOError:
            return Image.open("/Users/fedor/Desktop/traffic_signs/" + str(self.id) + ".png")

    def load_question(self):
        a = []
        if len(str(self.id)) == 1:
            temp = str(self.id) + " "
        else:
            temp = str(self.id)
        f = open(self.type + ".txt", "r")
        for x in f:
            if x[:2] == temp:
                a.append(unicode_to_ascii(x.rstrip('\n')))
                return (''.join(a)).split(";")
        f.close()

    def get_question(self):
        return self.question_data[1]

    def get_answer_options(self):
        answer_options = self.question_data[2].split(", ")
        random.shuffle(answer_options)
        return answer_options

    def get_answer(self):
        return self.question_data[3]

    # Method to compare two objects using == based on their id
    # Used to detect duplicates when deleting duplicated objects in a list of objects
    def __eq__(self, other):
        if isinstance(other, Question):
            if other.id == self.id:
                return True
        return False

    def answer_option_button_clicked(self, answer_guess):

        if answer_guess == self.get_answer():
            print("Well done! Correct")
            current_game_results.append('R' + str(self.id))
        else:
            print("Wrong. The correct answer was " + "'" + self.get_answer() + "'")
            current_game_results.append('W' + str(self.id))

        self.hide_question_frame()
        used_questions.append(self)

        if len(set_of_questions) == len(used_questions):
            save_game_result(current_game_results, self.type, self.name)
            current_user.set_up_menu_frame()

        for i in set_of_questions:
            if i not in used_questions:
                i.display_question_frame()
                break

    def display_question_frame(self):
        tk.Label(self.frame, text=self.get_question()).pack(pady=30)

        image1 = self.image
        width, height = image1.size
        image1 = image1.resize((width // 2, height // 2))
        test = ImageTk.PhotoImage(image1)
        label1 = tk.Label(self.frame, image=test)
        label1.image = test

        # Position image
        if self.type == "traffic_signs":
            label1.pack(pady=30)

        answer_options = self.get_answer_options()

        option1 = ttk.Button(
            self.frame,
            text=(answer_options[0]),
            command=lambda: self.answer_option_button_clicked(answer_options[0])
        )

        option2 = ttk.Button(
            self.frame,
            text=(answer_options[1]),
            command=lambda: self.answer_option_button_clicked(answer_options[1])
        )

        option3 = ttk.Button(
            self.frame,
            text=(answer_options[2]),
            command=lambda: self.answer_option_button_clicked(answer_options[2])
        )

        option4 = ttk.Button(
            self.frame,
            text=(answer_options[3]),
            command=lambda: self.answer_option_button_clicked(answer_options[3])
        )

        option1.pack(pady=3)
        option2.pack(pady=3)
        option3.pack(pady=3)
        option4.pack(pady=3)

        self.frame.pack(fill='both', expand=1)

    def hide_question_frame(self):
        self.frame.forget()


def create_set_of_questions(topic, number_of_questions, name):
    list_of_questions = []
    for i in range(number_of_questions):
        list_of_questions.append(Question(topic, name))
    temp = delete_duplicates(list_of_questions)
    while len(temp) != number_of_questions:
        for i in range(number_of_questions - len(temp)):
            list_of_questions.append(Question(topic, name))
        temp = delete_duplicates(list_of_questions)
    return temp
# Creates a list of Question objects with a specified number of them and topic (sent down as parameters)
# Questions are chosen at random and duplicates are deleted


def delete_duplicates(list_of_objects):
    temp = []
    for i in list_of_objects:
        if i not in temp:
            temp.append(i)
    return temp
# used to delete duplicates in the list of Question objects


def entrycheck(inp):
    if ' ' in inp or inp == '':  # don't accept spaces or empty inputs
        return False
    else:
        return True
# function to validate input for empty name and password when the user is registering


def restart_program():
    # Restarts the current program. Used as the logout option
    # Cleanup action (like saving data) is done before calling this function.
    python = sys.executable
    os.execl(python, python, * sys.argv)


root = tk.Tk()
root.title('Driving theory')
root.geometry("800x500")

image1 = Image.open("/Users/fedor/Downloads/logo1.png")
width, height = image1.size
image1 = image1.resize((width // 4, height // 4))
test = ImageTk.PhotoImage(image1)
label1 = tk.Label(root, image=test)
label1.pack(anchor=tk.NE)

tc = ttk.Notebook(root)

login = Authorisation("Login")
login.display_sign_in()
register = Authorisation("Register")
register.display_sign_in()

tc.add(login.frame, text=str(login.type))
tc.add(register.frame, text=str(register.type))
tc.pack(expand=1, fill="both")

tk.Button(root, text="Logout", command=restart_program).pack(side=tk.BOTTOM)

root.mainloop()
# Lines 757-871 is a driver code that sets up the application when program is executed
